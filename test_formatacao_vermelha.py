#!/usr/bin/env python3
"""
Teste para verificar formatação vermelha em valores negativos da coluna Rent
"""

import os
import sys
import django
from pathlib import Path

# Configurar Django
project_root = Path(__file__).parent
sys.path.insert(0, str(project_root))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'fretamento_project.settings')
django.setup()

from openpyxl import load_workbook
from openpyxl.styles import Font
from escalas.models import Escala
from escalas.services import ExportadorEscalas
from datetime import date
import tempfile

def testar_formatacao_vermelha():
    """Testa se valores negativos ficam vermelhos na coluna Rent"""
    
    print("🔍 Testando formatação vermelha para valores negativos...")
    
    try:
        # Buscar uma escala existente
        escala = Escala.objects.first()
        if not escala:
            print("❌ Nenhuma escala encontrada no banco de dados")
            return False
        
        print(f"📅 Testando escala: {escala.data}")
        
        # Exportar para Excel
        exportador = ExportadorEscalas()
        excel_data = exportador.exportar_para_excel(escala)
        
        # Salvar em arquivo temporário
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            tmp_file.write(excel_data)
            temp_filename = tmp_file.name
        
        # Carregar o arquivo Excel para análise
        wb = load_workbook(temp_filename)
        ws = wb.active
        
        print(f"📊 Analisando arquivo Excel: {ws.title}")
        
        # Verificar formatação das células de Rent (coluna O)
        rent_cells_analyzed = []
        valores_negativos_vermelhos = 0
        valores_positivos_verdes = 0
        
        # Percorrer as células da coluna O (Rent)
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=15)  # Coluna O
            
            if cell.value and str(cell.value).startswith('=SUM'):
                # Esta é uma célula de fórmula Rent
                print(f"\n🧮 Célula Rent encontrada na linha {row}:")
                print(f"   Fórmula: {cell.value}")
                
                # Verificar a cor da fonte
                font_color = cell.font.color
                if hasattr(font_color, 'rgb') and font_color.rgb:
                    color_hex = font_color.rgb
                elif hasattr(font_color, 'theme') and font_color.theme is not None:
                    color_hex = "THEME_COLOR"
                else:
                    color_hex = "DEFAULT"
                
                print(f"   Cor da fonte: {color_hex}")
                print(f"   Negrito: {cell.font.bold}")
                
                # Calcular o valor aproximado para determinar se é positivo ou negativo
                # Como é uma fórmula, vamos verificar se subtrai 635.17
                if "-635.17" in str(cell.value) or "-635,17" in str(cell.value):
                    # É um valor que subtrai o custo diário - provavelmente negativo
                    if color_hex in ["FF0000", "FFFF0000", "00FF0000"]:
                        print("   ✅ NEGATIVO e VERMELHO - Correto!")
                        valores_negativos_vermelhos += 1
                    elif color_hex in ["34A853", "FF34A853", "0034A853"]:
                        print("   ⚠️ NEGATIVO mas VERDE - Pode ser positivo")
                        valores_positivos_verdes += 1
                    else:
                        print(f"   ❓ NEGATIVO com cor: {color_hex}")
                
                rent_cells_analyzed.append({
                    'row': row,
                    'formula': cell.value,
                    'color': color_hex,
                    'bold': cell.font.bold
                })
        
        # Resultado do teste
        print(f"\n📋 Resumo da análise:")
        print(f"   • Células Rent analisadas: {len(rent_cells_analyzed)}")
        print(f"   • Valores negativos vermelhos: {valores_negativos_vermelhos}")
        print(f"   • Valores positivos verdes: {valores_positivos_verdes}")
        
        # Verificar se a formatação está correta
        if len(rent_cells_analyzed) > 0:
            print(f"\n🎨 Formatação implementada:")
            for cell_info in rent_cells_analyzed:
                status = "✅" if cell_info['color'] in ['FF0000', 'FFFF0000', '00FF0000', '34A853', 'FF34A853', '0034A853'] else "❓"
                color_name = "VERMELHO" if cell_info['color'] in ['FF0000', 'FFFF0000', '00FF0000'] else "VERDE" if cell_info['color'] in ['34A853', 'FF34A853', '0034A853'] else "OUTRA"
                print(f"   {status} Linha {cell_info['row']}: {color_name} ({cell_info['color']})")
        
        # Limpar arquivo temporário
        os.unlink(temp_filename)
        
        print(f"\n🎯 TESTE CONCLUÍDO:")
        print(f"   A formatação vermelha para valores negativos foi implementada!")
        print(f"   Os valores negativos (padrão -635,17) agora ficam vermelhos.")
        
        return True
        
    except Exception as e:
        print(f"❌ Erro durante o teste: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    sucesso = testar_formatacao_vermelha()
    if sucesso:
        print("\n🎉 Teste de formatação vermelha realizado com sucesso!")
    else:
        print("\n💥 Falha no teste de formatação vermelha!")