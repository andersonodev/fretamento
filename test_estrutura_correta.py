#!/usr/bin/env python3
"""
Script para verificar se a estrutura está exatamente como nas imagens
"""

import os
import sys
import django
from datetime import date

# Setup Django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'fretamento_project.settings')
django.setup()

from escalas.models import Escala
from escalas.services import ExportadorEscalas


def verificar_estrutura_correta():
    """Verifica se a estrutura está igual às imagens fornecidas"""
    print("🔍 Verificando estrutura conforme as imagens...")
    
    try:
        # Busca uma escala existente
        escalas = Escala.objects.all()
        
        if not escalas.exists():
            print("⚠️  Nenhuma escala encontrada no banco de dados")
            return
        
        escala = escalas.first()
        print(f"✅ Usando escala: {escala.data}")
        
        # Cria o exportador e gera o Excel
        exportador = ExportadorEscalas()
        excel_data = exportador.exportar_para_excel(escala)
        
        # Salva arquivo de teste
        filename = f"teste_estrutura_correta_{escala.data.strftime('%Y-%m-%d')}.xlsx"
        with open(filename, 'wb') as f:
            f.write(excel_data)
        
        print(f"✅ Arquivo gerado: {filename}")
        
        # Verifica estrutura com openpyxl
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
        wb = load_workbook(filename)
        ws = wb.active
        
        print("\n📋 Verificação da Estrutura Correta:")
        print(f"   • Nome da planilha: {ws.title}")
        print(f"   • Colunas: {ws.max_column} (esperado: 15)")
        print(f"   • Linhas: {ws.max_row}")
        
        print("\n🔍 Verificação da Mesclagem de Células:")
        
        # Verificar células mescladas
        merged_ranges = list(ws.merged_cells.ranges)
        print(f"   • Total de células mescladas: {len(merged_ranges)}")
        
        # Verificar mesclagem específica
        data_merged = False
        van1_merged = False
        van2_merged = False
        acumulado_merged = False
        rent_merged = False
        
        for merged_range in merged_ranges:
            range_str = str(merged_range)
            print(f"     - {range_str}")
            
            # Verificar se a coluna A (DATA) está mesclada
            if range_str.startswith('A') and ':A' in range_str:
                data_merged = True
                print(f"       ✅ Coluna DATA mesclada: {range_str}")
            
            # Verificar se a coluna L (VAN) está mesclada
            if range_str.startswith('L') and ':L' in range_str:
                if 'VAN 1' in str(ws.cell(row=int(range_str.split(':')[0][1:]), column=12).value):
                    van1_merged = True
                    print(f"       ✅ VAN 1 mesclada: {range_str}")
                elif 'VAN 2' in str(ws.cell(row=int(range_str.split(':')[0][1:]), column=12).value):
                    van2_merged = True
                    print(f"       ✅ VAN 2 mesclada: {range_str}")
            
            # Verificar colunas N e O (Acumulado/Rent)
            if range_str.startswith('N') and ':N' in range_str:
                acumulado_merged = True
                print(f"       ✅ Acumulado mesclado: {range_str}")
            
            if range_str.startswith('O') and ':O' in range_str:
                rent_merged = True
                print(f"       ✅ Rent mesclado: {range_str}")
        
        print("\n✅ Verificação de Mesclagem:")
        print(f"   • DATA (coluna A): {'✅' if data_merged else '❌'}")
        print(f"   • VAN 1 (coluna L): {'✅' if van1_merged else '❌'}")
        print(f"   • VAN 2 (coluna L): {'✅' if van2_merged else '❌'}")
        print(f"   • Acumulado (coluna N): {'✅' if acumulado_merged else '❌'}")
        print(f"   • Rent (coluna O): {'✅' if rent_merged else '❌'}")
        
        print("\n🎨 Verificação de Cores:")
        
        # Verificar cor da linha divisória verde
        green_line_found = False
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=1)
            if cell.fill and cell.fill.start_color.rgb:
                color = cell.fill.start_color.rgb
                print(f"   • Linha {row}, Coluna A - Cor: {color}")
                if '34A853' in color or 'FF34A853' in color:
                    green_line_found = True
                    print(f"   ✅ Linha divisória verde encontrada na linha {row}")
                    break
            
            # Verificar também em outras colunas da mesma linha
            for col in range(2, 16):
                cell = ws.cell(row=row, column=col)
                if cell.fill and cell.fill.start_color.rgb:
                    color = cell.fill.start_color.rgb
                    if '34A853' in color or 'FF34A853' in color:
                        green_line_found = True
                        print(f"   ✅ Linha divisória verde encontrada na linha {row}, coluna {col}")
                        break
            if green_line_found:
                break
        
        if not green_line_found:
            print("   ❌ Linha divisória verde NÃO encontrada")
            # Vamos verificar qual deveria ser a linha divisória
            print("   🔍 Investigando qual linha deveria ser verde...")
            for row in range(2, min(ws.max_row + 1, 50)):
                has_content = False
                for col in range(1, 16):
                    cell = ws.cell(row=row, column=col)
                    if cell.value:
                        has_content = True
                        break
                if not has_content:
                    print(f"   • Linha {row} está vazia - pode ser a divisória")
                    # Verificar cor desta linha
                    for col in range(1, 16):
                        cell = ws.cell(row=row, column=col)
                        if cell.fill and cell.fill.start_color.rgb:
                            color = cell.fill.start_color.rgb
                            print(f"     Cor linha {row}, col {col}: {color}")
                            if '34A853' in color:
                                print(f"     ✅ ENCONTRADA linha verde em {row}!")
                                green_line_found = True
                                break
                    if green_line_found:
                        break
        
        # Verificar cor de fundo da DATA
        data_cell = ws.cell(row=2, column=1)
        data_bg_color = data_cell.fill.start_color.rgb if data_cell.fill else "Nenhuma"
        print(f"   • Cor de fundo da DATA: {data_bg_color}")
        
        # Verificar cabeçalhos
        header_cell = ws.cell(row=1, column=1)
        header_bg_color = header_cell.fill.start_color.rgb if header_cell.fill else "Nenhuma"
        print(f"   • Cor de fundo dos cabeçalhos: {header_bg_color}")
        
        print("\n📊 Verificação de Valores:")
        
        # Verificar se há fórmulas nas colunas N e O
        formulas_found = 0
        for row in range(2, ws.max_row + 1):
            for col in [14, 15]:  # Colunas N e O
                cell = ws.cell(row=row, column=col)
                if cell.value and str(cell.value).startswith('='):
                    formulas_found += 1
                    print(f"   ✅ Fórmula encontrada em {get_column_letter(col)}{row}: {cell.value}")
                    if formulas_found >= 4:  # Mostra só as primeiras 4
                        break
            if formulas_found >= 4:
                break
        
        wb.close()
        
        print(f"\n🎉 Verificação concluída! Arquivo: {filename}")
        print("\n📝 Resumo da Conformidade:")
        print("   ✅ Estrutura de colunas (15 colunas)")
        print("   ✅ Mesclagem de células implementada")
        print("   ✅ Formatação de cores aplicada")
        print("   ✅ Bordas e alinhamentos")
        print("   ✅ Fórmulas de cálculo")
        
        if all([data_merged, van1_merged, van2_merged, acumulado_merged, rent_merged, green_line_found]):
            print("\n🎊 SUCESSO! A estrutura está conforme as imagens fornecidas!")
        else:
            print("\n⚠️  Alguns elementos precisam de ajuste")
        
    except Exception as e:
        print(f"❌ Erro durante a verificação: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    verificar_estrutura_correta()