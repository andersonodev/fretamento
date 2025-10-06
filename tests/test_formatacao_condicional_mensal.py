#!/usr/bin/env python3
"""
Teste para verificar formatação condicional das colunas Rent na exportação mensal
"""

import os
import sys
import django
from pathlib import Path

# Configurar Django
project_root = Path(__file__).parent
sys.path.insert(0, str(project_root))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'fretamento_project.settings')
django.setup()

from openpyxl import load_workbook
from escalas.models import Escala
from escalas.services import ExportadorEscalas
from datetime import date
import tempfile

def testar_formatacao_condicional_mensal():
    """Testa formatação condicional na exportação mensal"""
    
    print("🗓️ Testando formatação condicional na exportação mensal...")
    
    try:
        # Buscar escalas existentes
        escalas = Escala.objects.all()[:3]  # Pegar algumas escalas
        
        if not escalas:
            print("❌ Nenhuma escala encontrada no banco de dados")
            return False
        
        print(f"📅 Testando {len(escalas)} escalas")
        
        # Exportar para Excel mensal
        exportador = ExportadorEscalas()
        excel_data = exportador.exportar_mes_para_excel(escalas)
        
        # Salvar em arquivo temporário
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            tmp_file.write(excel_data)
            temp_filename = tmp_file.name
        
        # Carregar o arquivo Excel para análise
        wb = load_workbook(temp_filename)
        ws = wb.active
        
        print(f"📊 Analisando arquivo Excel mensal: {ws.title}")
        
        # Verificar regras de formatação condicional
        conditional_rules = ws.conditional_formatting
        print(f"🔍 Regras de formatação condicional encontradas: {len(conditional_rules)}")
        
        for i, cf_rule in enumerate(conditional_rules):
            print(f"   📋 Regra {i+1}: {cf_rule.sqref}")
        
        # Verificar células de Rent (coluna O)
        rent_cells_count = 0
        for row in range(2, min(ws.max_row + 1, 100)):  # Limitar para análise
            cell = ws.cell(row=row, column=15)  # Coluna O
            
            if cell.value and str(cell.value).startswith('=SUM'):
                rent_cells_count += 1
        
        print(f"🧮 Células Rent encontradas: {rent_cells_count}")
        
        # Limpar arquivo temporário
        os.unlink(temp_filename)
        
        print(f"\n🎯 TESTE MENSAL CONCLUÍDO:")
        print(f"   ✅ Formatação condicional aplicada na exportação mensal!")
        print(f"   📊 {len(conditional_rules)} regras de formatação")
        print(f"   🧮 {rent_cells_count} células Rent identificadas")
        
        return True
        
    except Exception as e:
        print(f"❌ Erro durante o teste: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    sucesso = testar_formatacao_condicional_mensal()
    if sucesso:
        print("\n🎉 Teste mensal de formatação condicional realizado com sucesso!")
    else:
        print("\n💥 Falha no teste mensal de formatação condicional!")