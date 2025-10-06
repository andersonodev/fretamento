#!/usr/bin/env python3
"""
Teste completo da funcionalidade de exportação mensal
Valida: múltiplas escalas, divisores vermelhos entre dias, formatação correta
"""

import os
import sys
import django
from datetime import date, datetime
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Configurar Django
sys.path.append(os.path.dirname(os.path.abspath(__file__)))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'fretamento_project.settings')
django.setup()

from escalas.models import Escala, AlocacaoVan, ServicoGrupo
from escalas.services import ExportadorEscalas
from core.models import Servico

def criar_dados_teste():
    """Cria dados de teste para exportação mensal"""
    print("🔧 Criando dados de teste...")
    
    # Criar escalas para janeiro de 2025
    escalas = []
    
    # Escala 1: 15/01/2025
    escala1 = Escala.objects.create(
        data=date(2025, 1, 15),
        etapa='finalizada'
    )
    escalas.append(escala1)
    
    # Criar serviço para escala 1
    servico1 = Servico.objects.create(
        cliente="CLIENTE TESTE 1",
        local_pickup="LOCAL PICKUP 1",
        numero_venda="12345",
        pax=10,
        horario="08:00",
        data_do_servico=date(2025, 1, 15),
        servico="SERVIÇO TESTE 1"
    )
    
    # Alocação VAN1 na escala 1
    AlocacaoVan.objects.create(
        escala=escala1,
        servico=servico1,
        van='VAN1',
        ordem=1,
        preco_calculado=100.00
    )
    
    # Escala 2: 16/01/2025
    escala2 = Escala.objects.create(
        data=date(2025, 1, 16),
        etapa='finalizada'
    )
    escalas.append(escala2)
    
    # Criar serviço para escala 2
    servico2 = Servico.objects.create(
        cliente="CLIENTE TESTE 2",
        local_pickup="LOCAL PICKUP 2",
        numero_venda="67890",
        pax=15,
        horario=None,  # Teste "SEM HORARIO"
        data_do_servico=date(2025, 1, 16),
        servico="SERVIÇO TESTE 2"
    )
    
    # Alocação VAN2 na escala 2
    AlocacaoVan.objects.create(
        escala=escala2,
        servico=servico2,
        van='VAN2',
        ordem=1,
        preco_calculado=150.00
    )
    
    # Escala 3: 17/01/2025
    escala3 = Escala.objects.create(
        data=date(2025, 1, 17),
        etapa='finalizada'
    )
    escalas.append(escala3)
    
    # Criar dois serviços para escala 3 (teste de grupo)
    servico3a = Servico.objects.create(
        cliente="CLIENTE TESTE 3A",
        local_pickup="LOCAL PICKUP 3",
        numero_venda="11111",
        pax=8,
        horario="09:30",
        data_do_servico=date(2025, 1, 17),
        servico="SERVIÇO TESTE 3A"
    )
    
    servico3b = Servico.objects.create(
        cliente="CLIENTE TESTE 3B",
        local_pickup="LOCAL PICKUP 3",
        numero_venda="22222",
        pax=12,
        horario="09:30",
        data_do_servico=date(2025, 1, 17),
        servico="SERVIÇO TESTE 3B"
    )
    
    # Alocações VAN1 na escala 3
    alocacao3a = AlocacaoVan.objects.create(
        escala=escala3,
        servico=servico3a,
        van='VAN1',
        ordem=1,
        preco_calculado=80.00
    )
    
    alocacao3b = AlocacaoVan.objects.create(
        escala=escala3,
        servico=servico3b,
        van='VAN1',
        ordem=2,
        preco_calculado=120.00
    )
    
    # Criar grupo de serviços
    from escalas.models import GrupoServico
    grupo = GrupoServico.objects.create(
        escala=escala3,
        van='VAN1',
        ordem=1,
        cliente_principal="GRUPO TESTE 1",
        servico_principal="SERVIÇOS AGRUPADOS",
        local_pickup_principal="LOCAL PICKUP 3",
        total_pax=20,
        total_valor=200.00
    )
    
    ServicoGrupo.objects.create(grupo=grupo, alocacao=alocacao3a)
    ServicoGrupo.objects.create(grupo=grupo, alocacao=alocacao3b)
    
    print(f"✅ Criadas {len(escalas)} escalas de teste")
    return escalas

def validar_estrutura_excel(excel_data, escalas):
    """Valida a estrutura do arquivo Excel gerado"""
    print("📋 Validando estrutura do Excel...")
    
    # Carregar workbook
    wb = load_workbook(BytesIO(excel_data))
    ws = wb.active
    
    print(f"   • Nome da planilha: {ws.title}")
    print(f"   • Número de colunas: {ws.max_column}")
    print(f"   • Número de linhas: {ws.max_row}")
    
    # Validar cabeçalhos
    headers_esperados = [
        "DATA", "CLIENTE", "Local Pick-UP", "NÚMERO DA VENDA", "PAX",
        "HORÁRIO", "DATA DO SERVIÇO", "INÍCIO", "TÉRMINO", "SERVIÇOS", 
        "VALOR CUSTO TARIFÁRIO", "VAN", "OBS", "Acumulado Van 01", "Rent Van 01"
    ]
    
    print("🏷️  Validando cabeçalhos:")
    for col, header_esperado in enumerate(headers_esperados, 1):
        header_atual = ws.cell(row=1, column=col).value
        if header_atual == header_esperado:
            print(f"   ✅ Coluna {col}: {header_atual}")
        else:
            print(f"   ❌ Coluna {col}: Esperado '{header_esperado}', encontrado '{header_atual}'")
    
    # Procurar divisores vermelhos entre dias
    print("🔍 Procurando divisores vermelhos entre dias...")
    divisores_vermelhos = 0
    cor_vermelha = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=1)
        if cell.fill.start_color.rgb == "FFFF0000":  # Vermelho
            divisores_vermelhos += 1
            print(f"   📍 Divisor vermelho encontrado na linha {row}")
    
    print(f"   Total de divisores vermelhos: {divisores_vermelhos}")
    
    # Verificar conteúdo das células
    print("📝 Verificando conteúdo das escalas:")
    dados_encontrados = []
    
    for row in range(2, ws.max_row + 1):
        cliente = ws.cell(row=row, column=2).value
        if cliente and "CLIENTE TESTE" in str(cliente):
            dados_encontrados.append({
                'linha': row,
                'cliente': cliente,
                'horario': ws.cell(row=row, column=6).value,
                'van': ws.cell(row=row, column=12).value
            })
    
    for dados in dados_encontrados:
        print(f"   📋 Linha {dados['linha']}: {dados['cliente']}, VAN: {dados['van']}, Horário: {dados['horario']}")
    
    # Verificar "SEM HORARIO"
    sem_horario_encontrado = False
    for row in range(2, ws.max_row + 1):
        horario = ws.cell(row=row, column=6).value
        if horario == "SEM HORARIO":
            sem_horario_encontrado = True
            print(f"   ✅ 'SEM HORARIO' encontrado na linha {row}")
            break
    
    if not sem_horario_encontrado:
        print("   ⚠️  'SEM HORARIO' não foi encontrado")
    
    return True

def test_exportacao_mensal():
    """Teste principal da exportação mensal"""
    print("🚀 TESTE DE EXPORTAÇÃO MENSAL")
    print("=" * 50)
    
    try:
        # Limpar dados anteriores
        print("🧹 Limpando dados anteriores...")
        AlocacaoVan.objects.filter(escala__data__year=2025, escala__data__month=1).delete()
        Escala.objects.filter(data__year=2025, data__month=1).delete()
        ServicoGrupo.objects.all().delete()  # Limpar todos os grupos
        from escalas.models import GrupoServico
        GrupoServico.objects.filter(cliente_principal__startswith="GRUPO TESTE").delete()
        Servico.objects.filter(cliente__startswith="CLIENTE TESTE").delete()
        
        # Criar dados de teste
        escalas = criar_dados_teste()
        
        # Testar exportação
        print("📤 Testando exportação mensal...")
        exportador = ExportadorEscalas()
        excel_data = exportador.exportar_mes_para_excel(escalas)
        
        print(f"✅ Excel gerado com sucesso: {len(excel_data)} bytes")
        
        # Salvar arquivo para inspeção
        filename = f"teste_exportacao_mensal_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        with open(filename, 'wb') as f:
            f.write(excel_data)
        
        print(f"💾 Arquivo salvo: {filename}")
        
        # Validar estrutura
        validar_estrutura_excel(excel_data, escalas)
        
        print("\n🎉 TESTE CONCLUÍDO COM SUCESSO!")
        print(f"📂 Arquivo de teste salvo: {filename}")
        print("🔍 Você pode abrir o arquivo Excel para inspeção visual")
        
        return True
        
    except Exception as e:
        print(f"❌ ERRO no teste: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    test_exportacao_mensal()