#!/usr/bin/env python3
"""
Script para testar o novo formato de exportação Excel
"""

import os
import sys
import django
from datetime import date

# Setup Django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'fretamento_project.settings')
django.setup()

from escalas.models import Escala
from escalas.services import ExportadorEscalas


def test_export_format():
    """Testa o formato de exportação Excel"""
    print("🧪 Testando novo formato de exportação Excel...")
    
    try:
        # Busca uma escala existente ou usa uma data padrão
        escalas = Escala.objects.all()
        
        if escalas.exists():
            escala = escalas.first()
            print(f"✅ Usando escala existente: {escala.data}")
        else:
            print("⚠️  Nenhuma escala encontrada no banco de dados")
            print("   Para testar, crie uma escala primeiro através da interface web")
            return
        
        # Cria o exportador
        exportador = ExportadorEscalas()
        
        # Gera o Excel
        print("📊 Gerando arquivo Excel...")
        excel_data = exportador.exportar_para_excel(escala)
        
        # Salva arquivo de teste
        filename = f"teste_formato_excel_{escala.data.strftime('%Y-%m-%d')}.xlsx"
        with open(filename, 'wb') as f:
            f.write(excel_data)
        
        print(f"✅ Arquivo gerado com sucesso: {filename}")
        print(f"📂 Tamanho: {len(excel_data)} bytes")
        
        # Verifica estrutura com openpyxl
        from openpyxl import load_workbook
        wb = load_workbook(filename)
        ws = wb.active
        
        print("\n📋 Estrutura do arquivo:")
        print(f"   • Nome da planilha: {ws.title}")
        print(f"   • Número de colunas: {ws.max_column}")
        print(f"   • Número de linhas: {ws.max_row}")
        
        print("\n🏷️  Cabeçalhos das colunas:")
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            print(f"   • Coluna {col}: {header}")
        
        # Verifica se tem dados
        data_rows = ws.max_row - 1  # Exclui cabeçalho
        print(f"\n📊 Linhas de dados: {data_rows}")
        
        if data_rows > 0:
            print("✅ Arquivo contém dados da escala")
        else:
            print("⚠️  Arquivo não contém dados (escala vazia)")
        
        wb.close()
        
        print(f"\n🎉 Teste concluído! Arquivo salvo como: {filename}")
        print("   Abra o arquivo para verificar se está no formato desejado")
        
    except Exception as e:
        print(f"❌ Erro durante o teste: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    test_export_format()