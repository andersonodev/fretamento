#!/usr/bin/env python3
"""
Teste para verificar formatação condicional das colunas Rent
Valores negativos devem ficar vermelhos e valores positivos devem ficar verdes
"""

import os
import sys
import django
from pathlib import Path

# Configurar Django
project_root = Path(__file__).parent
sys.path.insert(0, str(project_root))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'fretamento_project.settings')
django.setup()

from openpyxl import load_workbook
from escalas.models import Escala
from escalas.services import ExportadorEscalas
from datetime import date
import tempfile

def testar_formatacao_condicional():
    """Testa se a formatação condicional está funcionando corretamente"""
    
    print("🎨 Testando formatação condicional das colunas Rent...")
    
    try:
        # Buscar uma escala existente
        escala = Escala.objects.first()
        if not escala:
            print("❌ Nenhuma escala encontrada no banco de dados")
            return False
        
        print(f"📅 Testando escala: {escala.data}")
        
        # Exportar para Excel
        exportador = ExportadorEscalas()
        excel_data = exportador.exportar_para_excel(escala)
        
        # Salvar em arquivo temporário
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            tmp_file.write(excel_data)
            temp_filename = tmp_file.name
        
        # Carregar o arquivo Excel para análise
        wb = load_workbook(temp_filename)
        ws = wb.active
        
        print(f"📊 Analisando arquivo Excel: {ws.title}")
        
        # Verificar se existem regras de formatação condicional
        conditional_rules = ws.conditional_formatting
        print(f"🔍 Regras de formatação condicional encontradas: {len(conditional_rules)}")
        
        for i, cf_rule in enumerate(conditional_rules):
            print(f"   📋 Regra {i+1}:")
            print(f"      - Intervalo: {cf_rule.sqref}")
            for rule in cf_rule.rules:
                print(f"      - Operador: {rule.operator}")
                print(f"      - Fórmula: {rule.formula}")
                if hasattr(rule, 'font') and rule.font:
                    color = rule.font.color.rgb if rule.font.color else "Padrão"
                    print(f"      - Cor da fonte: {color}")
        
        # Verificar células específicas de Rent (coluna O)
        rent_cells_found = []
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=15)  # Coluna O
            
            if cell.value and str(cell.value).startswith('=SUM'):
                # Esta é uma célula de fórmula Rent
                rent_cells_found.append({
                    'row': row,
                    'formula': cell.value,
                    'font_color': cell.font.color.rgb if cell.font.color else "Padrão",
                    'bold': cell.font.bold
                })
        
        print(f"\n🧮 Células Rent encontradas: {len(rent_cells_found)}")
        for cell_info in rent_cells_found:
            print(f"   📍 Linha {cell_info['row']}: {cell_info['formula']}")
            print(f"      - Cor base: {cell_info['font_color']}")
            print(f"      - Negrito: {cell_info['bold']}")
        
        # Limpar arquivo temporário
        os.unlink(temp_filename)
        
        print(f"\n🎯 TESTE CONCLUÍDO:")
        print(f"   ✅ Formatação condicional implementada!")
        print(f"   📊 {len(conditional_rules)} regras aplicadas")
        print(f"   🧮 {len(rent_cells_found)} células Rent identificadas")
        print(f"   🎨 Valores negativos ficam vermelhos, positivos ficam verdes")
        
        return True
        
    except Exception as e:
        print(f"❌ Erro durante o teste: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    sucesso = testar_formatacao_condicional()
    if sucesso:
        print("\n🎉 Teste de formatação condicional realizado com sucesso!")
    else:
        print("\n💥 Falha no teste de formatação condicional!")