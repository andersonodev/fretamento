"""
Serviços para gerenciamento de escalas
"""

from datetime import datetime, date
from typing import List, Dict
from django.db import transaction
from core.models import Servico, GrupoServico
from escalas.models import Escala, AlocacaoVan, ServicoGrupo
from core.logic import OtimizadorEscalas, CalculadorVeiculoPreco
import logging

logger = logging.getLogger(__name__)


class GerenciadorEscalas:
    """Classe principal para gerenciar escalas - DEPRECATED"""
    
    def __init__(self):
        # Mantido para compatibilidade, mas não usado mais
        pass
    
    def criar_escala_basica(self, data_alvo: date) -> 'Escala':
        """
        Cria apenas a estrutura básica da escala (sem dados)
        """
        from escalas.models import Escala
        
        # Busca ou cria a escala
        escala, created = Escala.objects.get_or_create(
            data=data_alvo,
            defaults={'etapa': 'ESTRUTURA'}
        )
        
        if created:
            logger.info(f"Estrutura de escala criada para {data_alvo}")
        else:
            logger.info(f"Escala para {data_alvo} já existe.")
        
        return escala


class ExportadorEscalas:
    """Classe para exportar escalas em diversos formatos"""
    
    def exportar_para_excel(self, escala: Escala) -> bytes:
        """Exporta escala para formato Excel conforme especificação visual completa"""
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from django.db.models import Q
        
        wb = Workbook()
        ws = wb.active
        nomeMes = escala.data.strftime('%B')
        ws.title = f"Escala para {nomeMes}"
        
        # Cabeçalhos conforme especificação
        headers = [
            "DATA", "CLIENTE", "Local Pick-UP", "NÚMERO DA VENDA", "PAX",
            "HORÁRIO", "DATA DO SERVIÇO", "INÍCIO", "TÉRMINO", "SERVIÇOS", 
            "VALOR CUSTO TARIFÁRIO", "VAN", "OBS", "Acumulado Van 01", "Rent Van 01"
        ]
        
        # ===== ESTILOS E FORMATAÇÃO =====
        # Cabeçalhos: fundo cinza-claro, negrito, centralizado, bordas
        header_font = Font(bold=True, size=10)
        header_fill = PatternFill(start_color="E5E5E5", end_color="E5E5E5", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Data: fundo verde-claro, negrito
        data_font = Font(bold=True, size=10)
        data_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
        data_alignment = Alignment(horizontal="left", vertical="center")
        
        # Linha divisória verde
        divisor_fill = PatternFill(start_color="34A853", end_color="34A853", fill_type="solid")
        
        # Van resumo: fundo cinza claro
        van_resumo_fill = PatternFill(start_color="F3F3F3", end_color="F3F3F3", fill_type="solid")
        van_resumo_font = Font(bold=True)
        center_alignment = Alignment(horizontal="center", vertical="center")
        right_alignment = Alignment(horizontal="right", vertical="center")
        left_alignment = Alignment(horizontal="left", vertical="center")
        
        # Formatação condicional para Rent
        rent_negative_font = Font(color="FF0000", bold=True)  # Vermelho
        rent_positive_font = Font(color="00B050", bold=True)  # Verde
        
        # Escreve cabeçalhos
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Configurar larguras de coluna
        column_widths = [80, 110, 120, 110, 65, 65, 80, 65, 65, 300, 120, 70, 150, 120, 120]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width / 7
        
        # Congelar primeira linha
        ws.freeze_panes = "A2"
        
        # Função para processar grupos e serviços individuais
        def processar_van(alocacoes, van_name):
            """Processa alocações de uma van, agrupando quando necessário"""
            grupos_processados = set()
            resultados = []
            
            for alocacao in alocacoes:
                # Verificar se esta alocação está em um grupo
                try:
                    grupo_info = alocacao.grupo_info
                    grupo = grupo_info.grupo
                    
                    if grupo.id not in grupos_processados:
                        # Processar grupo completo
                        grupos_processados.add(grupo.id)
                        
                        # Buscar todos os serviços do grupo na mesma van
                        servicos_do_grupo = []
                        alocacoes_do_grupo = alocacao.escala.alocacoes.filter(
                            van=alocacao.van,
                            grupo_info__grupo=grupo
                        ).order_by('ordem')
                        
                        for aloc in alocacoes_do_grupo:
                            servicos_do_grupo.append(aloc)
                        
                        if len(servicos_do_grupo) > 1:
                            # Grupo com múltiplos serviços - concatenar números de venda
                            numeros_venda = []
                            total_pax = 0
                            primeiro_servico = servicos_do_grupo[0].servico
                            
                            for aloc_grupo in servicos_do_grupo:
                                if aloc_grupo.servico.numero_venda:
                                    numeros_venda.append(str(aloc_grupo.servico.numero_venda))
                                total_pax += aloc_grupo.servico.pax or 0
                            
                            # Criar linha agrupada
                            linha_grupo = {
                                'data': escala.data,
                                'cliente': primeiro_servico.cliente,
                                'local_pickup': primeiro_servico.local_pickup or "",
                                'numero_venda': " / ".join(numeros_venda),  # CONCATENADO
                                'pax': total_pax,
                                'horario': primeiro_servico.horario,
                                'data_servico': primeiro_servico.data_do_servico,
                                'servico': f"{primeiro_servico.servico} (+{len(servicos_do_grupo)-1} / Grupo)",
                                'preco': sum(float(aloc.preco_calculado or 0) for aloc in servicos_do_grupo),
                                'van': van_name,
                                'obs': f"Grupo {grupo.id}"
                            }
                            resultados.append(linha_grupo)
                        else:
                            # Grupo com apenas um serviço - tratar como individual
                            resultados.append({
                                'data': escala.data,
                                'cliente': alocacao.servico.cliente,
                                'local_pickup': alocacao.servico.local_pickup or "",
                                'numero_venda': alocacao.servico.numero_venda or "",
                                'pax': alocacao.servico.pax,
                                'horario': alocacao.servico.horario,
                                'data_servico': alocacao.servico.data_do_servico,
                                'servico': alocacao.servico.servico,
                                'preco': float(alocacao.preco_calculado or 0),
                                'van': van_name,
                                'obs': ""
                            })
                
                except ServicoGrupo.DoesNotExist:
                    # Serviço individual (não está em grupo)
                    resultados.append({
                        'data': escala.data,
                        'cliente': alocacao.servico.cliente,
                        'local_pickup': alocacao.servico.local_pickup or "",
                        'numero_venda': alocacao.servico.numero_venda or "",
                        'pax': alocacao.servico.pax,
                        'horario': alocacao.servico.horario,
                        'data_servico': alocacao.servico.data_do_servico,
                        'servico': alocacao.servico.servico,
                        'preco': float(alocacao.preco_calculado or 0),
                        'van': van_name,
                        'obs': ""
                    })
            
            return resultados
        
        # Obtém alocações da escala ordenadas por van e ordem
        alocacoes_van1 = escala.alocacoes.filter(van='VAN1').order_by('ordem')
        alocacoes_van2 = escala.alocacoes.filter(van='VAN2').order_by('ordem')
        
        # Processar cada van
        dados_van1 = processar_van(alocacoes_van1, "VAN 1")
        dados_van2 = processar_van(alocacoes_van2, "VAN 2")
        
        row = 2
        CUSTO_DIARIO = 635.17
        
        # ===== LINHA DE DATA (CABEÇALHO VISUAL) =====
        # Primeira linha com a data em destaque
        data_cell = ws.cell(row=row, column=1, value=escala.data.strftime('%d/%m/%Y'))
        data_cell.font = data_font
        data_cell.fill = data_fill
        data_cell.alignment = data_alignment
        data_cell.border = thin_border
        
        # Mesclar a linha da data por várias colunas para efeito visual
        ws.merge_cells(f"A{row}:F{row}")
        for col in range(1, 7):
            ws.cell(row=row, column=col).fill = data_fill
            ws.cell(row=row, column=col).border = thin_border
        
        row += 1
        
        # ===== VAN 1 =====
        van1_start_row = row
        van1_rows = 0
        
        for dados in dados_van1:
            # Preencher dados da linha
            ws.cell(row=row, column=1, value="")  # DATA (vazia nas linhas de dados)
            ws.cell(row=row, column=2, value=dados['cliente'])  # CLIENTE
            ws.cell(row=row, column=3, value=dados['local_pickup'])  # Local Pick-UP
            ws.cell(row=row, column=4, value=dados['numero_venda'])  # NÚMERO DA VENDA
            ws.cell(row=row, column=5, value=dados['pax'])  # PAX
            ws.cell(row=row, column=6, value=dados['horario'])  # HORÁRIO
            ws.cell(row=row, column=7, value=dados['data_servico'])  # DATA DO SERVIÇO
            ws.cell(row=row, column=8, value="")  # INÍCIO (vazio)
            ws.cell(row=row, column=9, value="")  # TÉRMINO (vazio)
            ws.cell(row=row, column=10, value=dados['servico'])  # SERVIÇOS
            ws.cell(row=row, column=11, value="")  # VALOR CUSTO TARIFÁRIO (vazio)
            ws.cell(row=row, column=12, value=dados['van'])  # VAN
            ws.cell(row=row, column=13, value=dados['obs'])  # OBS
            
            # Aplicar bordas e alinhamentos
            for col in range(1, 16):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                if col in [4, 5, 11, 14, 15]:  # Números e valores à direita
                    cell.alignment = right_alignment
                else:
                    cell.alignment = left_alignment
            
            row += 1
            van1_rows += 1
        
        # Preencher linhas vazias para Van 1 (mínimo 20 linhas)
        MIN_ROWS_VAN = 20
        while van1_rows < MIN_ROWS_VAN:
            for col in range(1, 16):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                if col == 12:
                    cell.value = "VAN 1"
                    cell.alignment = center_alignment
            row += 1
            van1_rows += 1
        
        van1_end_row = row - 1
        
        # ===== LINHA DE RESUMO VAN 1 =====
        # Calcular total da Van 1
        total_van1 = sum(dados['preco'] for dados in dados_van1)
        rent_van1 = total_van1 - CUSTO_DIARIO
        
        # Linha de resumo Van 1
        resumo_row = row
        ws.cell(row=resumo_row, column=14, value=f"R$ {total_van1:.2f}".replace('.', ','))  # Acumulado
        ws.cell(row=resumo_row, column=15, value=f"–R$ {abs(rent_van1):.2f}".replace('.', ','))  # Rent
        
        # Formatação da linha de resumo
        for col in [14, 15]:
            cell = ws.cell(row=resumo_row, column=col)
            cell.fill = van_resumo_fill
            cell.font = van_resumo_font
            cell.alignment = center_alignment
            cell.border = thin_border
            
        # Formatação condicional do Rent
        rent_cell = ws.cell(row=resumo_row, column=15)
        if rent_van1 < -CUSTO_DIARIO:
            rent_cell.font = rent_negative_font  # Vermelho
        else:
            rent_cell.font = rent_positive_font  # Verde
        
        row += 1
        
        # ===== LINHA DIVISÓRIA VERDE =====
        for col in range(1, 16):
            cell = ws.cell(row=row, column=col)
            cell.fill = divisor_fill
            cell.border = thin_border
        
        row += 1
        
        # ===== VAN 2 =====
        van2_start_row = row
        van2_rows = 0
        
        for dados in dados_van2:
            # Preencher dados da linha
            ws.cell(row=row, column=1, value="")  # DATA (vazia)
            ws.cell(row=row, column=2, value=dados['cliente'])  # CLIENTE
            ws.cell(row=row, column=3, value=dados['local_pickup'])  # Local Pick-UP
            ws.cell(row=row, column=4, value=dados['numero_venda'])  # NÚMERO DA VENDA
            ws.cell(row=row, column=5, value=dados['pax'])  # PAX
            ws.cell(row=row, column=6, value=dados['horario'])  # HORÁRIO
            ws.cell(row=row, column=7, value=dados['data_servico'])  # DATA DO SERVIÇO
            ws.cell(row=row, column=8, value="")  # INÍCIO (vazio)
            ws.cell(row=row, column=9, value="")  # TÉRMINO (vazio)
            ws.cell(row=row, column=10, value=dados['servico'])  # SERVIÇOS
            ws.cell(row=row, column=11, value="")  # VALOR CUSTO TARIFÁRIO (vazio)
            ws.cell(row=row, column=12, value=dados['van'])  # VAN
            ws.cell(row=row, column=13, value=dados['obs'])  # OBS
            
            # Aplicar bordas e alinhamentos
            for col in range(1, 16):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                if col in [4, 5, 11, 14, 15]:  # Números e valores à direita
                    cell.alignment = right_alignment
                else:
                    cell.alignment = left_alignment
            
            row += 1
            van2_rows += 1
        
        # Preencher linhas vazias para Van 2 (mínimo 20 linhas)
        while van2_rows < MIN_ROWS_VAN:
            for col in range(1, 16):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                if col == 12:
                    cell.value = "VAN 2"
                    cell.alignment = center_alignment
            row += 1
            van2_rows += 1
        
        van2_end_row = row - 1
        
        # ===== LINHA DE RESUMO VAN 2 =====
        if van2_rows > 0:
            # Calcular total da Van 2
            total_van2 = sum(dados['preco'] for dados in dados_van2)
            rent_van2 = total_van2 - CUSTO_DIARIO
            
            # Linha de resumo Van 2
            resumo_row = row
            ws.cell(row=resumo_row, column=14, value=f"R$ {total_van2:.2f}".replace('.', ','))  # Acumulado
            ws.cell(row=resumo_row, column=15, value=f"–R$ {abs(rent_van2):.2f}".replace('.', ','))  # Rent
            
            # Formatação da linha de resumo
            for col in [14, 15]:
                cell = ws.cell(row=resumo_row, column=col)
                cell.fill = van_resumo_fill
                cell.font = van_resumo_font
                cell.alignment = center_alignment
                cell.border = thin_border
            
            # Formatação condicional do Rent
            rent_cell = ws.cell(row=resumo_row, column=15)
            if rent_van2 < -CUSTO_DIARIO:
                rent_cell.font = rent_negative_font  # Vermelho
            else:
                rent_cell.font = rent_positive_font  # Verde
        
        # ===== FORMATAÇÃO FINAL =====
        # Formato de data
        for row_num in range(2, row + 1):
            ws.cell(row=row_num, column=7).number_format = 'dd/mm/yyyy'  # DATA DO SERVIÇO
        
        # Formato de horário
        for row_num in range(2, row + 1):
            ws.cell(row=row_num, column=6).number_format = 'hh:mm'  # HORÁRIO
            ws.cell(row=row_num, column=8).number_format = 'hh:mm'  # INÍCIO
            ws.cell(row=row_num, column=9).number_format = 'hh:mm'  # TÉRMINO
        
        # Salva em buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer.getvalue()