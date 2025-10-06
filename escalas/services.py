"""
Serviços para gerenciamento de escalas
"""

from datetime import datetime, date
from typing import List, Dict
from django.db import transaction
from core.models import Servico, GrupoServico
from escalas.models import Escala, AlocacaoVan, ServicoGrupo
from core.logic import OtimizadorEscalas, CalculadorVeiculoPreco
import logging

logger = logging.getLogger(__name__)


class GerenciadorEscalas:
    """Classe principal para gerenciar escalas - DEPRECATED"""
    
    def __init__(self):
        # Mantido para compatibilidade, mas não usado mais
        pass
    
    def criar_escala_basica(self, data_alvo: date) -> 'Escala':
        """
        Cria apenas a estrutura básica da escala (sem dados)
        """
        from escalas.models import Escala
        
        # Busca ou cria a escala
        escala, created = Escala.objects.get_or_create(
            data=data_alvo,
            defaults={'etapa': 'ESTRUTURA'}
        )
        
        if created:
            logger.info(f"Estrutura de escala criada para {data_alvo}")
        else:
            logger.info(f"Escala para {data_alvo} já existe.")
        
        return escala


class ExportadorEscalas:
    """Classe para exportar escalas em diversos formatos"""
    
    def exportar_para_excel(self, escala: Escala) -> bytes:
        """Exporta escala para formato Excel seguindo exatamente a estrutura das imagens"""
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from django.db.models import Q
        
        wb = Workbook()
        ws = wb.active
        nomeMes = escala.data.strftime('%B')
        ws.title = f"Escala para {nomeMes}"
        
        # Cabeçalhos conforme o código Google Apps Script
        headers = [
            "DATA", "CLIENTE", "Local Pick-UP", "NÚMERO DA VENDA", "PAX",
            "HORÁRIO", "DATA DO SERVIÇO", "INÍCIO", "TÉRMINO", "SERVIÇOS", 
            "VALOR CUSTO TARIFÁRIO", "VAN", "OBS", "Acumulado Van 01", "Rent Van 01"
        ]
        
        # ===== ESTILOS E FORMATAÇÃO =====
        # Cabeçalhos: fundo verde-claro, negrito, centralizado
        header_font = Font(bold=True, size=10)
        header_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # DATA e VAN: fundo cinza, negrito, centralizado
        data_van_font = Font(bold=True, size=10)
        data_van_fill = PatternFill(start_color="EFEFEF", end_color="EFEFEF", fill_type="solid")
        center_alignment = Alignment(horizontal="center", vertical="center")
        
        # Linha divisória verde
        divisor_fill = PatternFill(start_color="34A853", end_color="34A853", fill_type="solid")
        
        # Resumo: fundo cinza, negrito, centralizado
        resumo_font = Font(bold=True, size=10)
        resumo_fill = PatternFill(start_color="EFEFEF", end_color="EFEFEF", fill_type="solid")
        
        # Formatação condicional para Rent
        rent_positive_font = Font(color="34A853", bold=True)  # Verde para positivo
        rent_negative_font = Font(color="FF0000", bold=True)  # Vermelho para negativo
        
        left_alignment = Alignment(horizontal="left", vertical="center")
        right_alignment = Alignment(horizontal="right", vertical="center")
        
        # Escreve cabeçalhos
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Configurar larguras de coluna conforme Google Apps Script
        column_widths = [80, 110, 120, 110, 65, 65, 80, 65, 65, 300, 120, 70, 150, 120, 120]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width / 7
        
        # Congelar primeira linha
        ws.freeze_panes = "A2"
        
        # Função para processar grupos e serviços individuais
        def processar_van(alocacoes, van_name):
            """Processa alocações de uma van, agrupando quando necessário"""
            grupos_processados = set()
            resultados = []
            
            for alocacao in alocacoes:
                # Verificar se esta alocação está em um grupo
                try:
                    grupo_info = alocacao.grupo_info
                    grupo = grupo_info.grupo
                    
                    if grupo.id not in grupos_processados:
                        # Processar grupo completo
                        grupos_processados.add(grupo.id)
                        
                        # Buscar todos os serviços do grupo na mesma van
                        servicos_do_grupo = []
                        alocacoes_do_grupo = alocacao.escala.alocacoes.filter(
                            van=alocacao.van,
                            grupo_info__grupo=grupo
                        ).order_by('ordem')
                        
                        for aloc in alocacoes_do_grupo:
                            servicos_do_grupo.append(aloc)
                        
                        if len(servicos_do_grupo) > 1:
                            # Grupo com múltiplos serviços - concatenar números de venda
                            numeros_venda = []
                            total_pax = 0
                            primeiro_servico = servicos_do_grupo[0].servico
                            
                            for aloc_grupo in servicos_do_grupo:
                                if aloc_grupo.servico.numero_venda:
                                    numeros_venda.append(str(aloc_grupo.servico.numero_venda))
                                total_pax += aloc_grupo.servico.pax or 0
                            
                            # Criar linha agrupada
                            linha_grupo = {
                                'cliente': primeiro_servico.cliente,
                                'local_pickup': primeiro_servico.local_pickup or "",
                                'numero_venda': " / ".join(numeros_venda),
                                'pax': total_pax,
                                'horario': primeiro_servico.horario if primeiro_servico.horario else "SEM HORARIO",
                                'data_servico': primeiro_servico.data_do_servico,
                                'servico': f"{primeiro_servico.servico} (+{len(servicos_do_grupo)-1} / Grupo)",
                                'preco': sum(float(aloc.preco_calculado or 0) for aloc in servicos_do_grupo),
                                'obs': f"Grupo {grupo.id}"
                            }
                            resultados.append(linha_grupo)
                        else:
                            # Grupo com apenas um serviço - tratar como individual
                            resultados.append({
                                'cliente': alocacao.servico.cliente,
                                'local_pickup': alocacao.servico.local_pickup or "",
                                'numero_venda': alocacao.servico.numero_venda or "",
                                'pax': alocacao.servico.pax,
                                'horario': alocacao.servico.horario if alocacao.servico.horario else "SEM HORARIO",
                                'data_servico': alocacao.servico.data_do_servico,
                                'servico': alocacao.servico.servico,
                                'preco': float(alocacao.preco_calculado or 0),
                                'obs': ""
                            })
                
                except ServicoGrupo.DoesNotExist:
                    # Serviço individual (não está em grupo)
                    resultados.append({
                        'cliente': alocacao.servico.cliente,
                        'local_pickup': alocacao.servico.local_pickup or "",
                        'numero_venda': alocacao.servico.numero_venda or "",
                        'pax': alocacao.servico.pax,
                        'horario': alocacao.servico.horario if alocacao.servico.horario else "SEM HORARIO",
                        'data_servico': alocacao.servico.data_do_servico,
                        'servico': alocacao.servico.servico,
                        'preco': float(alocacao.preco_calculado or 0),
                        'obs': ""
                    })
            
            return resultados
        
        # Obtém alocações da escala ordenadas por van e ordem
        alocacoes_van1 = escala.alocacoes.filter(van='VAN1').order_by('ordem')
        alocacoes_van2 = escala.alocacoes.filter(van='VAN2').order_by('ordem')
        
        # Processar cada van
        dados_van1 = processar_van(alocacoes_van1, "VAN 1")
        dados_van2 = processar_van(alocacoes_van2, "VAN 2")
        
        # Constantes
        CUSTO_DIARIO = 635.17
        MIN_ROWS_VAN = 20
        
        row = 2
        
        # ===== CALCULAR DIMENSÕES DO BLOCO =====
        altura_van1 = max(len(dados_van1), MIN_ROWS_VAN)
        altura_van2 = max(len(dados_van2), MIN_ROWS_VAN)
        altura_total_bloco = altura_van1 + altura_van2 + 1  # +1 para linha divisória
        
        van1_start_row = row
        van1_end_row = row + altura_van1 - 1
        divisor_row = row + altura_van1
        van2_start_row = row + altura_van1 + 1
        van2_end_row = row + altura_van1 + altura_van2
        
        # ===== APLICAR BORDAS EM TODO O BLOCO =====
        for r in range(van1_start_row, van2_end_row + 1):
            for c in range(1, 16):
                cell = ws.cell(row=r, column=c)
                cell.border = thin_border
        
        # ===== MESCLAR E PREENCHER COLUNA DATA (A) =====
        ws.merge_cells(f"A{van1_start_row}:A{van2_end_row}")
        data_cell = ws.cell(row=van1_start_row, column=1)
        data_cell.value = escala.data
        data_cell.font = data_van_font
        data_cell.fill = data_van_fill
        data_cell.alignment = center_alignment
        data_cell.number_format = 'dd/mm/yy'
        
        # ===== MESCLAR E PREENCHER VAN 1 (L) =====
        ws.merge_cells(f"L{van1_start_row}:L{van1_end_row}")
        van1_cell = ws.cell(row=van1_start_row, column=12)
        van1_cell.value = "VAN 1"
        van1_cell.font = data_van_font
        van1_cell.fill = data_van_fill
        van1_cell.alignment = center_alignment
        
        # ===== PREENCHER DADOS VAN 1 =====
        current_row = van1_start_row
        for dados in dados_van1:
            ws.cell(row=current_row, column=2, value=dados['cliente'])
            ws.cell(row=current_row, column=3, value=dados['local_pickup'])
            ws.cell(row=current_row, column=4, value=dados['numero_venda'])
            ws.cell(row=current_row, column=5, value=dados['pax'])
            ws.cell(row=current_row, column=6, value=dados['horario'])
            ws.cell(row=current_row, column=7, value=dados['data_servico'])
            ws.cell(row=current_row, column=8, value="")  # INÍCIO
            ws.cell(row=current_row, column=9, value="")  # TÉRMINO
            ws.cell(row=current_row, column=10, value=dados['servico'])
            ws.cell(row=current_row, column=11, value=dados['preco'])
            ws.cell(row=current_row, column=13, value=dados['obs'])
            
            current_row += 1
        
        # ===== MESCLAR E PREENCHER ACUMULADO/RENT VAN 1 (N/O) =====
        total_van1 = sum(dados['preco'] for dados in dados_van1)
        rent_van1 = total_van1 - CUSTO_DIARIO
        
        ws.merge_cells(f"N{van1_start_row}:N{van1_end_row}")
        acumulado1_cell = ws.cell(row=van1_start_row, column=14)
        acumulado1_cell.value = f"=SUM(K{van1_start_row}:K{van1_end_row})"
        acumulado1_cell.font = resumo_font
        acumulado1_cell.fill = resumo_fill
        acumulado1_cell.alignment = center_alignment
        acumulado1_cell.number_format = 'R$ #,##0.00'
        
        ws.merge_cells(f"O{van1_start_row}:O{van1_end_row}")
        rent1_cell = ws.cell(row=van1_start_row, column=15)
        rent1_cell.value = f"=SUM(K{van1_start_row}:K{van1_end_row})-{CUSTO_DIARIO}"
        # Aplicar cor vermelha por padrão para valores que subtraem o custo diário
        # (assumindo que normalmente resultará em valor negativo)
        rent1_cell.font = rent_negative_font  # Sempre vermelho para -635.17
        rent1_cell.fill = resumo_fill
        rent1_cell.alignment = center_alignment
        rent1_cell.number_format = 'R$ #,##0.00'
        
        # ===== LINHA DIVISÓRIA VERDE =====
        for col in range(1, 16):
            cell = ws.cell(row=divisor_row, column=col)
            cell.fill = divisor_fill
            cell.border = thin_border
        
        # ===== MESCLAR E PREENCHER VAN 2 (L) =====
        ws.merge_cells(f"L{van2_start_row}:L{van2_end_row}")
        van2_cell = ws.cell(row=van2_start_row, column=12)
        van2_cell.value = "VAN 2"
        van2_cell.font = data_van_font
        van2_cell.fill = data_van_fill
        van2_cell.alignment = center_alignment
        
        # ===== PREENCHER DADOS VAN 2 =====
        current_row = van2_start_row
        for dados in dados_van2:
            ws.cell(row=current_row, column=2, value=dados['cliente'])
            ws.cell(row=current_row, column=3, value=dados['local_pickup'])
            ws.cell(row=current_row, column=4, value=dados['numero_venda'])
            ws.cell(row=current_row, column=5, value=dados['pax'])
            ws.cell(row=current_row, column=6, value=dados['horario'])
            ws.cell(row=current_row, column=7, value=dados['data_servico'])
            ws.cell(row=current_row, column=8, value="")  # INÍCIO
            ws.cell(row=current_row, column=9, value="")  # TÉRMINO
            ws.cell(row=current_row, column=10, value=dados['servico'])
            ws.cell(row=current_row, column=11, value=dados['preco'])
            ws.cell(row=current_row, column=13, value=dados['obs'])
            
            current_row += 1
        
        # ===== MESCLAR E PREENCHER ACUMULADO/RENT VAN 2 (N/O) =====
        total_van2 = sum(dados['preco'] for dados in dados_van2)
        rent_van2 = total_van2 - CUSTO_DIARIO
        
        ws.merge_cells(f"N{van2_start_row}:N{van2_end_row}")
        acumulado2_cell = ws.cell(row=van2_start_row, column=14)
        acumulado2_cell.value = f"=SUM(K{van2_start_row}:K{van2_end_row})"
        acumulado2_cell.font = resumo_font
        acumulado2_cell.fill = resumo_fill
        acumulado2_cell.alignment = center_alignment
        acumulado2_cell.number_format = 'R$ #,##0.00'
        
        ws.merge_cells(f"O{van2_start_row}:O{van2_end_row}")
        rent2_cell = ws.cell(row=van2_start_row, column=15)
        rent2_cell.value = f"=SUM(K{van2_start_row}:K{van2_end_row})-{CUSTO_DIARIO}"
        # Aplicar cor vermelha por padrão para valores que subtraem o custo diário
        # (assumindo que normalmente resultará em valor negativo)
        rent2_cell.font = rent_negative_font  # Sempre vermelho para -635.17
        rent2_cell.fill = resumo_fill
        rent2_cell.alignment = center_alignment
        rent2_cell.number_format = 'R$ #,##0.00'
        
        # ===== FORMATAÇÃO FINAL =====
        # Formato de moeda para coluna K
        for r in range(van1_start_row, van2_end_row + 1):
            ws.cell(row=r, column=11).number_format = 'R$ #,##0.00'
        
        # Formato de horário
        for r in range(van1_start_row, van2_end_row + 1):
            ws.cell(row=r, column=6).number_format = 'hh:mm'  # HORÁRIO
            ws.cell(row=r, column=8).number_format = 'hh:mm'  # INÍCIO
            ws.cell(row=r, column=9).number_format = 'hh:mm'  # TÉRMINO
        
        # Formato de data
        for r in range(van1_start_row, van2_end_row + 1):
            ws.cell(row=r, column=7).number_format = 'dd/mm/yyyy'  # DATA DO SERVIÇO
        
        # Salva em buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer.getvalue()

    def exportar_mes_para_excel(self, escalas_mes: list) -> bytes:
        """Exporta múltiplas escalas do mês para formato Excel com divisores vermelhos entre dias"""
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from django.db.models import Q
        
        if not escalas_mes:
            # Se não há escalas, retorna um Excel vazio
            wb = Workbook()
            ws = wb.active
            ws.title = "Escalas Mensais"
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return buffer.getvalue()

        wb = Workbook()
        ws = wb.active
        primeiro_mes = escalas_mes[0].data.strftime('%B %Y')
        ws.title = f"Escalas {primeiro_mes}"

        # Mesmo cabeçalho do método diário
        headers = [
            "DATA", "CLIENTE", "Local Pick-UP", "NÚMERO DA VENDA", "PAX",
            "HORÁRIO", "DATA DO SERVIÇO", "INÍCIO", "TÉRMINO", "SERVIÇOS", 
            "VALOR CUSTO TARIFÁRIO", "VAN", "OBS", "Acumulado Van 01", "Rent Van 01"
        ]

        # ===== ESTILOS E FORMATAÇÃO (mesmos do método diário) =====
        header_font = Font(bold=True, size=10)
        header_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # Data/Van: fundo verde-claro, negrito
        data_van_font = Font(bold=True, size=10)
        data_van_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")

        # Linha divisória verde (entre VAN1 e VAN2 no mesmo dia)
        divisor_fill = PatternFill(start_color="34A853", end_color="34A853", fill_type="solid")
        
        # NOVO: Linha divisória vermelha (entre dias diferentes)
        divisor_dia_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

        # Resumo: fundo cinza-claro
        resumo_fill = PatternFill(start_color="F3F3F3", end_color="F3F3F3", fill_type="solid")
        resumo_font = Font(bold=True)

        # Formatação condicional para Rent
        rent_positive_font = Font(color="34A853", bold=True)  # Verde para positivo
        rent_negative_font = Font(color="FF0000", bold=True)  # Vermelho para negativo

        left_alignment = Alignment(horizontal="left", vertical="center")
        right_alignment = Alignment(horizontal="right", vertical="center")
        center_alignment = Alignment(horizontal="center", vertical="center")

        # Escreve cabeçalhos
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Configurar larguras de coluna (mesmas do método diário)
        column_widths = [12, 25, 25, 15, 8, 12, 15, 8, 8, 30, 15, 10, 20, 15, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # Constantes
        CUSTO_DIARIO = 635.17
        MIN_ROWS_VAN = 20
        
        current_row = 2
        
        # Processa cada escala (dia) do mês
        for i, escala in enumerate(escalas_mes):
            # Se não é o primeiro dia, adiciona divisor vermelho
            if i > 0:
                for col in range(1, 16):
                    cell = ws.cell(row=current_row, column=col)
                    cell.fill = divisor_dia_fill
                    cell.border = thin_border
                current_row += 1

            # Processa dados da escala atual (mesmo código do método diário)
            def processar_van(alocacoes, van_name):
                """Processa alocações de uma van, agrupando quando necessário"""
                resultados = []
                servicos_processados = set()
                
                for alocacao in alocacoes:
                    # Evita processar o mesmo serviço múltiplas vezes se estiver em grupo
                    if alocacao.servico.id in servicos_processados:
                        continue
                    
                    try:
                        # Verifica se a alocação está em um grupo através do relacionamento grupo_info
                        grupo_servico = alocacao.grupo_info
                        grupo = grupo_servico.grupo
                        
                        # Busca todas as alocações do mesmo grupo na mesma escala
                        alocacoes_do_grupo = AlocacaoVan.objects.filter(
                            escala=alocacao.escala,
                            van=alocacao.van,
                            grupo_info__grupo=grupo
                        ).order_by('ordem')
                        
                        servicos_do_grupo = []
                        for aloc in alocacoes_do_grupo:
                            servicos_do_grupo.append(aloc)
                        
                        if len(servicos_do_grupo) > 1:
                            # Grupo com múltiplos serviços - concatenar números de venda
                            numeros_venda = []
                            total_pax = 0
                            primeiro_servico = servicos_do_grupo[0].servico
                            
                            for aloc_grupo in servicos_do_grupo:
                                if aloc_grupo.servico.numero_venda:
                                    numeros_venda.append(str(aloc_grupo.servico.numero_venda))
                                total_pax += aloc_grupo.servico.pax or 0
                            
                            # Criar linha agrupada
                            linha_grupo = {
                                'cliente': primeiro_servico.cliente,
                                'local_pickup': primeiro_servico.local_pickup or "",
                                'numero_venda': " / ".join(numeros_venda),
                                'pax': total_pax,
                                'horario': primeiro_servico.horario if primeiro_servico.horario else "SEM HORARIO",
                                'data_servico': primeiro_servico.data_do_servico,
                                'servico': f"GRUPO #{grupo.id} - {primeiro_servico.servico}",
                                'preco': sum(float(aloc.preco_calculado or 0) for aloc in servicos_do_grupo),
                                'obs': ""
                            }
                            resultados.append(linha_grupo)
                            
                            # Marca todos os serviços do grupo como processados
                            for aloc_grupo in servicos_do_grupo:
                                servicos_processados.add(aloc_grupo.servico.id)
                        else:
                            # Grupo com apenas um serviço
                            resultados.append({
                                'cliente': alocacao.servico.cliente,
                                'local_pickup': alocacao.servico.local_pickup or "",
                                'numero_venda': alocacao.servico.numero_venda or "",
                                'pax': alocacao.servico.pax,
                                'horario': alocacao.servico.horario if alocacao.servico.horario else "SEM HORARIO",
                                'data_servico': alocacao.servico.data_do_servico,
                                'servico': alocacao.servico.servico,
                                'preco': float(alocacao.preco_calculado or 0),
                                'obs': ""
                            })
                            servicos_processados.add(alocacao.servico.id)
                
                    except (AttributeError, ServicoGrupo.DoesNotExist):
                        # Serviço individual (não está em grupo)
                        resultados.append({
                            'cliente': alocacao.servico.cliente,
                            'local_pickup': alocacao.servico.local_pickup or "",
                            'numero_venda': alocacao.servico.numero_venda or "",
                            'pax': alocacao.servico.pax,
                            'horario': alocacao.servico.horario if alocacao.servico.horario else "SEM HORARIO",
                            'data_servico': alocacao.servico.data_do_servico,
                            'servico': alocacao.servico.servico,
                            'preco': float(alocacao.preco_calculado or 0),
                            'obs': ""
                        })
                        servicos_processados.add(alocacao.servico.id)
                
                return resultados

            # Processa VAN1 e VAN2 da escala atual
            alocacoes_van1 = escala.alocacoes.filter(van='VAN1').order_by('ordem')
            alocacoes_van2 = escala.alocacoes.filter(van='VAN2').order_by('ordem')
            
            dados_van1 = processar_van(alocacoes_van1, "VAN 1")
            dados_van2 = processar_van(alocacoes_van2, "VAN 2")

            # ===== CALCULAR DIMENSÕES DO BLOCO PARA ESTE DIA =====
            total_rows_van1 = max(len(dados_van1), MIN_ROWS_VAN)
            total_rows_van2 = max(len(dados_van2), MIN_ROWS_VAN)
            
            van1_start_row = current_row
            van1_end_row = van1_start_row + total_rows_van1 - 1
            divisor_row = van1_end_row + 1
            van2_start_row = divisor_row + 1
            van2_end_row = van2_start_row + total_rows_van2 - 1

            # ===== APLICAR BORDAS PARA TODO O BLOCO =====
            for r in range(van1_start_row, van2_end_row + 1):
                for c in range(1, 16):
                    cell = ws.cell(row=r, column=c)
                    cell.border = thin_border

            # ===== MESCLAR E PREENCHER COLUNA DATA (A) =====
            ws.merge_cells(f"A{van1_start_row}:A{van2_end_row}")
            data_cell = ws.cell(row=van1_start_row, column=1)
            data_cell.value = escala.data
            data_cell.font = data_van_font
            data_cell.fill = data_van_fill
            data_cell.alignment = center_alignment
            data_cell.number_format = 'dd/mm/yy'

            # ===== MESCLAR E PREENCHER VAN 1 (L) =====
            ws.merge_cells(f"L{van1_start_row}:L{van1_end_row}")
            van1_cell = ws.cell(row=van1_start_row, column=12)
            van1_cell.value = "VAN 1"
            van1_cell.font = data_van_font
            van1_cell.fill = data_van_fill
            van1_cell.alignment = center_alignment

            # ===== PREENCHER DADOS VAN 1 =====
            current_van_row = van1_start_row
            for dados in dados_van1:
                ws.cell(row=current_van_row, column=2, value=dados['cliente'])
                ws.cell(row=current_van_row, column=3, value=dados['local_pickup'])
                ws.cell(row=current_van_row, column=4, value=dados['numero_venda'])
                ws.cell(row=current_van_row, column=5, value=dados['pax'])
                ws.cell(row=current_van_row, column=6, value=dados['horario'])
                ws.cell(row=current_van_row, column=7, value=dados['data_servico'])
                ws.cell(row=current_van_row, column=8, value="")  # INÍCIO
                ws.cell(row=current_van_row, column=9, value="")  # TÉRMINO
                ws.cell(row=current_van_row, column=10, value=dados['servico'])
                ws.cell(row=current_van_row, column=11, value=dados['preco'])
                ws.cell(row=current_van_row, column=13, value=dados['obs'])
                current_van_row += 1

            # ===== MESCLAR E PREENCHER ACUMULADO/RENT VAN 1 (N/O) =====
            ws.merge_cells(f"N{van1_start_row}:N{van1_end_row}")
            acumulado1_cell = ws.cell(row=van1_start_row, column=14)
            acumulado1_cell.value = f"=SUM(K{van1_start_row}:K{van1_end_row})"
            acumulado1_cell.font = resumo_font
            acumulado1_cell.fill = resumo_fill
            acumulado1_cell.alignment = center_alignment
            acumulado1_cell.number_format = 'R$ #,##0.00'

            ws.merge_cells(f"O{van1_start_row}:O{van1_end_row}")
            rent1_cell = ws.cell(row=van1_start_row, column=15)
            rent1_cell.value = f"=SUM(K{van1_start_row}:K{van1_end_row})-{CUSTO_DIARIO}"
            rent1_cell.font = rent_negative_font  # Sempre vermelho para -635.17
            rent1_cell.fill = resumo_fill
            rent1_cell.alignment = center_alignment
            rent1_cell.number_format = 'R$ #,##0.00'

            # ===== LINHA DIVISÓRIA VERDE (entre VAN1 e VAN2 do mesmo dia) =====
            for col in range(1, 16):
                cell = ws.cell(row=divisor_row, column=col)
                cell.fill = divisor_fill
                cell.border = thin_border

            # ===== MESCLAR E PREENCHER VAN 2 (L) =====
            ws.merge_cells(f"L{van2_start_row}:L{van2_end_row}")
            van2_cell = ws.cell(row=van2_start_row, column=12)
            van2_cell.value = "VAN 2"
            van2_cell.font = data_van_font
            van2_cell.fill = data_van_fill
            van2_cell.alignment = center_alignment

            # ===== PREENCHER DADOS VAN 2 =====
            current_van_row = van2_start_row
            for dados in dados_van2:
                ws.cell(row=current_van_row, column=2, value=dados['cliente'])
                ws.cell(row=current_van_row, column=3, value=dados['local_pickup'])
                ws.cell(row=current_van_row, column=4, value=dados['numero_venda'])
                ws.cell(row=current_van_row, column=5, value=dados['pax'])
                ws.cell(row=current_van_row, column=6, value=dados['horario'])
                ws.cell(row=current_van_row, column=7, value=dados['data_servico'])
                ws.cell(row=current_van_row, column=8, value="")  # INÍCIO
                ws.cell(row=current_van_row, column=9, value="")  # TÉRMINO
                ws.cell(row=current_van_row, column=10, value=dados['servico'])
                ws.cell(row=current_van_row, column=11, value=dados['preco'])
                ws.cell(row=current_van_row, column=13, value=dados['obs'])
                current_van_row += 1

            # ===== MESCLAR E PREENCHER ACUMULADO/RENT VAN 2 (N/O) =====
            ws.merge_cells(f"N{van2_start_row}:N{van2_end_row}")
            acumulado2_cell = ws.cell(row=van2_start_row, column=14)
            acumulado2_cell.value = f"=SUM(K{van2_start_row}:K{van2_end_row})"
            acumulado2_cell.font = resumo_font
            acumulado2_cell.fill = resumo_fill
            acumulado2_cell.alignment = center_alignment
            acumulado2_cell.number_format = 'R$ #,##0.00'

            ws.merge_cells(f"O{van2_start_row}:O{van2_end_row}")
            rent2_cell = ws.cell(row=van2_start_row, column=15)
            rent2_cell.value = f"=SUM(K{van2_start_row}:K{van2_end_row})-{CUSTO_DIARIO}"
            rent2_cell.font = rent_negative_font  # Sempre vermelho para -635.17
            rent2_cell.fill = resumo_fill
            rent2_cell.alignment = center_alignment
            rent2_cell.number_format = 'R$ #,##0.00'

            # ===== APLICAR FORMATAÇÃO DE NÚMERO/DATA =====
            # Formato de moeda (coluna K - VALOR CUSTO TARIFÁRIO)
            for r in range(van1_start_row, van2_end_row + 1):
                ws.cell(row=r, column=11).number_format = 'R$ #,##0.00'

            # Formato de hora
            for r in range(van1_start_row, van2_end_row + 1):
                ws.cell(row=r, column=6).number_format = 'hh:mm'  # HORÁRIO
                ws.cell(row=r, column=8).number_format = 'hh:mm'  # INÍCIO
                ws.cell(row=r, column=9).number_format = 'hh:mm'  # TÉRMINO

            # Formato de data
            for r in range(van1_start_row, van2_end_row + 1):
                ws.cell(row=r, column=7).number_format = 'dd/mm/yyyy'  # DATA DO SERVIÇO

            # Atualiza current_row para próxima escala
            current_row = van2_end_row + 1

        # Salva em buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer.getvalue()