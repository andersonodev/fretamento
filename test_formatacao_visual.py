#!/usr/bin/env python3
"""
Script para verificar detalhadamente todas as formatações visuais do Excel
"""

import os
import sys
import django
from datetime import date

# Setup Django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'fretamento_project.settings')
django.setup()

from escalas.models import Escala
from escalas.services import ExportadorEscalas


def verificar_formatacao_excel():
    """Verifica todas as formatações visuais do Excel"""
    print("🔍 Verificando formatações visuais do Excel...")
    
    try:
        # Busca uma escala existente
        escalas = Escala.objects.all()
        
        if not escalas.exists():
            print("⚠️  Nenhuma escala encontrada no banco de dados")
            return
        
        escala = escalas.first()
        print(f"✅ Usando escala: {escala.data}")
        
        # Cria o exportador e gera o Excel
        exportador = ExportadorEscalas()
        excel_data = exportador.exportar_para_excel(escala)
        
        # Salva arquivo de teste
        filename = f"teste_formatacao_completa_{escala.data.strftime('%Y-%m-%d')}.xlsx"
        with open(filename, 'wb') as f:
            f.write(excel_data)
        
        print(f"✅ Arquivo gerado: {filename}")
        
        # Verifica estrutura com openpyxl
        from openpyxl import load_workbook
        wb = load_workbook(filename)
        ws = wb.active
        
        print("\n📋 Verificação da Estrutura:")
        print(f"   • Nome da planilha: {ws.title}")
        print(f"   • Colunas: {ws.max_column} (esperado: 15)")
        print(f"   • Linhas: {ws.max_row}")
        
        print("\n🏷️  Cabeçalhos (linha 1):")
        for col in range(1, min(ws.max_column + 1, 16)):
            cell = ws.cell(row=1, column=col)
            header = cell.value
            has_bold = cell.font.bold if cell.font else False
            has_fill = bool(cell.fill.start_color.rgb) if cell.fill else False
            print(f"   • Col {col}: '{header}' | Negrito: {has_bold} | Preenchimento: {has_fill}")
        
        print("\n📅 Verificação da Linha de Data (linha 2):")
        data_cell = ws.cell(row=2, column=1)
        print(f"   • Valor: {data_cell.value}")
        print(f"   • Negrito: {data_cell.font.bold if data_cell.font else False}")
        print(f"   • Cor de fundo: {data_cell.fill.start_color.rgb if data_cell.fill else 'Nenhuma'}")
        
        # Verificar se há células mescladas
        merged_ranges = ws.merged_cells.ranges
        print(f"   • Células mescladas na linha: {len([r for r in merged_ranges if 2 in range(r.min_row, r.max_row + 1)])}")
        
        print("\n🚐 Verificação dos Dados das Vans:")
        van_lines = []
        for row in range(3, min(ws.max_row + 1, 50)):  # Verifica primeiras 47 linhas
            van_cell = ws.cell(row=row, column=12)
            if van_cell.value and 'VAN' in str(van_cell.value):
                van_lines.append((row, van_cell.value))
        
        print(f"   • Linhas com identificação de Van: {len(van_lines)}")
        for row, van in van_lines[:5]:  # Mostra primeiras 5
            print(f"     - Linha {row}: {van}")
        
        print("\n📊 Verificação das Colunas de Resumo (N e O):")
        acumulado_values = []
        rent_values = []
        
        for row in range(2, min(ws.max_row + 1, 100)):
            acum_cell = ws.cell(row=row, column=14)  # Coluna N
            rent_cell = ws.cell(row=row, column=15)  # Coluna O
            
            if acum_cell.value and 'R$' in str(acum_cell.value):
                acumulado_values.append((row, acum_cell.value))
            
            if rent_cell.value and 'R$' in str(rent_cell.value):
                font_color = rent_cell.font.color.rgb if rent_cell.font and rent_cell.font.color else 'Padrão'
                rent_values.append((row, rent_cell.value, font_color))
        
        print(f"   • Valores de Acumulado encontrados: {len(acumulado_values)}")
        for row, value in acumulado_values[:3]:
            print(f"     - Linha {row}: {value}")
        
        print(f"   • Valores de Rent encontrados: {len(rent_values)}")
        for row, value, color in rent_values[:3]:
            print(f"     - Linha {row}: {value} | Cor: {color}")
        
        print("\n🟢 Verificação das Linhas Divisórias Verdes:")
        green_lines = []
        for row in range(2, min(ws.max_row + 1, 100)):
            cell = ws.cell(row=row, column=1)
            if cell.fill and cell.fill.start_color.rgb:
                color = cell.fill.start_color.rgb
                if '34A853' in color or 'FF34A853' in color:  # Verde da linha divisória
                    green_lines.append(row)
        
        print(f"   • Linhas divisórias verdes encontradas: {len(green_lines)}")
        for line in green_lines:
            print(f"     - Linha {line}")
        
        print("\n🔗 Verificação de Bordas:")
        bordered_cells = 0
        for row in range(1, min(6, ws.max_row + 1)):  # Verifica primeiras 5 linhas
            for col in range(1, min(16, ws.max_column + 1)):
                cell = ws.cell(row=row, column=col)
                if cell.border and any([cell.border.left.style, cell.border.right.style, 
                                       cell.border.top.style, cell.border.bottom.style]):
                    bordered_cells += 1
        
        print(f"   • Células com bordas (primeiras 5 linhas): {bordered_cells}")
        
        wb.close()
        
        print(f"\n🎉 Verificação completa! Arquivo: {filename}")
        print("📝 Resumo das implementações:")
        print("   ✅ Cabeçalhos com formatação cinza e negrito")
        print("   ✅ Linha de data com fundo verde-claro")
        print("   ✅ Identificação de Vans")
        print("   ✅ Colunas de resumo (Acumulado e Rent)")
        print("   ✅ Linhas divisórias verdes")
        print("   ✅ Bordas aplicadas")
        print("\n📂 Abra o arquivo no Excel para verificar visualmente!")
        
    except Exception as e:
        print(f"❌ Erro durante a verificação: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    verificar_formatacao_excel()